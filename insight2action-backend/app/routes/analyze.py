import io
import json
import csv
import logging
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from typing import Optional
from app.schemas.analysis import (
    AnalyzeRequest, AnalyzeResponse,
    RegenerateActionRequest, RegenerateActionResponse
)
from app.services.agent_orchestrator import run_agentic_workflow, run_regenerate_action

router = APIRouter()
logger = logging.getLogger(__name__)

# Supported file extensions and MIME types
SUPPORTED_EXTENSIONS = {
    '.txt', '.csv', '.json', '.md', '.log', '.xml', '.html', '.htm',
    '.tsv', '.yaml', '.yml', '.ini', '.cfg', '.conf', '.py', '.js',
    '.ts', '.jsx', '.tsx', '.sql', '.r', '.sas', '.stata',
}
BINARY_EXTENSIONS = {'.xlsx', '.xls', '.docx', '.doc', '.pdf', '.pptx', '.ppt'}
ALL_SUPPORTED = SUPPORTED_EXTENSIONS | BINARY_EXTENSIONS

MAX_FILE_SIZE = 2 * 1024 * 1024  # 2MB

def extract_text_from_csv(raw_bytes: bytes) -> str:
    """Convert CSV bytes to a readable text table."""
    try:
        text = raw_bytes.decode('utf-8', errors='replace')
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return text
        # Format as readable table
        lines = []
        for i, row in enumerate(rows[:200]):  # Cap at 200 rows
            lines.append(" | ".join(row))
        if len(rows) > 200:
            lines.append(f"... and {len(rows) - 200} more rows")
        return "\n".join(lines)
    except Exception:
        return raw_bytes.decode('utf-8', errors='replace')

def extract_text_from_xlsx(raw_bytes: bytes) -> str:
    """Extract text from XLSX using openpyxl if available, else basic ZIP parsing."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                all_text.append(" | ".join(cells))
        wb.close()
        return "\n".join(all_text)
    except ImportError:
        # Fallback: try to extract from shared strings XML inside XLSX zip
        import zipfile
        try:
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
                if 'xl/sharedStrings.xml' in z.namelist():
                    data = z.read('xl/sharedStrings.xml').decode('utf-8', errors='replace')
                    import re
                    texts = re.findall(r'<t[^>]*>(.*?)</t>', data)
                    return "\n".join(texts[:500])
            return "[Could not parse XLSX file. Install openpyxl for full support.]"
        except Exception:
            return "[Could not parse XLSX file.]"

def extract_text_from_docx(raw_bytes: bytes) -> str:
    """Extract text from DOCX by reading XML paragraphs from the ZIP."""
    import zipfile
    import re
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
            if 'word/document.xml' in z.namelist():
                xml_data = z.read('word/document.xml').decode('utf-8', errors='replace')
                # Extract text between <w:t> tags
                texts = re.findall(r'<w:t[^>]*>(.*?)</w:t>', xml_data)
                # Group by paragraphs (rough heuristic)
                paragraphs = []
                current = []
                for t in texts:
                    current.append(t)
                paragraphs.append(" ".join(current))
                return "\n".join(paragraphs)
        return "[Could not extract text from DOCX.]"
    except Exception:
        return "[Could not extract text from DOCX.]"

def extract_text_from_pdf(raw_bytes: bytes) -> str:
    """Extract text from PDF using PyPDF2/pypdf if available."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw_bytes))
        text_parts = []
        for page in reader.pages[:50]:  # Cap at 50 pages
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        if text_parts:
            return "\n\n".join(text_parts)
        return "[PDF contained no extractable text. It may be image-based.]"
    except ImportError:
        return "[PDF support requires pypdf. Install with: pip install pypdf]"
    except Exception as e:
        return f"[Failed to parse PDF: {str(e)}]"

@router.post("/analyze", response_model=AnalyzeResponse)
async def analyze_content(request: AnalyzeRequest):
    """
    Analyzes unstructured content and returns a structured insight and simulated action.
    """
    response_data = run_agentic_workflow(request.content)
    return response_data

@router.post("/analyze-file", response_model=AnalyzeResponse)
async def analyze_file(
    file: UploadFile = File(...),
):
    """
    Accepts file uploads of various formats, extracts text, and runs analysis.
    Supported: TXT, CSV, JSON, XLSX, DOCX, PDF, MD, LOG, XML, HTML, and more.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided.")
    
    # Check file extension
    import os
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALL_SUPPORTED:
        raise HTTPException(
            status_code=400, 
            detail=f"Unsupported file type: {ext}. Supported: {', '.join(sorted(ALL_SUPPORTED))}"
        )
    
    # Read file bytes
    raw_bytes = await file.read()
    if len(raw_bytes) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 2MB.")
    if len(raw_bytes) == 0:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    
    # Extract text based on type
    try:
        if ext == '.csv':
            content = extract_text_from_csv(raw_bytes)
        elif ext in {'.xlsx', '.xls'}:
            content = extract_text_from_xlsx(raw_bytes)
        elif ext in {'.docx', '.doc'}:
            content = extract_text_from_docx(raw_bytes)
        elif ext == '.pdf':
            content = extract_text_from_pdf(raw_bytes)
        elif ext == '.json':
            try:
                parsed = json.loads(raw_bytes.decode('utf-8', errors='replace'))
                content = json.dumps(parsed, indent=2)[:10000]  # Cap at 10k chars
            except json.JSONDecodeError:
                content = raw_bytes.decode('utf-8', errors='replace')
        else:
            # All other text-based files
            content = raw_bytes.decode('utf-8', errors='replace')
    except Exception as e:
        logger.error(f"Error extracting text from {file.filename}: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")
    
    if not content.strip():
        raise HTTPException(status_code=400, detail="No text could be extracted from the file.")
    
    # Add file context to help the AI
    enriched_content = f"[File: {file.filename} | Type: {ext}]\n\n{content}"
    
    response_data = run_agentic_workflow(enriched_content)
    return response_data

@router.post("/regenerate-action", response_model=RegenerateActionResponse)
async def regenerate_action(request: RegenerateActionRequest):
    """
    Regenerates the recommended action and alternatives based on user feedback.
    """
    try:
        response_data = run_regenerate_action(request.analysis_id, request.feedback)
        return response_data
    except Exception as e:
        logger.error(f"Error regenerating action: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to regenerate action: {str(e)}")
